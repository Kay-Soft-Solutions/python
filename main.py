from kivy.app import App
from kivy.uix.button import Button
from kivy.uix.label import Label
from kivy.uix.textinput import TextInput
from kivy.uix.boxlayout import BoxLayout
from kivy.graphics import Color, Rectangle
from openpyxl import load_workbook

class MyWidgetContainer(BoxLayout):
    def __init__(self, **kwargs):
        super(MyWidgetContainer, self).__init__(orientation='vertical', padding=10, spacing=10, **kwargs)
        
        with self.canvas.before:
            Color(0.0, 0.3, 0.2, 1)  # Set the background color (R, G, B, A)
            self.rect = Rectangle(size=self.size, pos=self.pos)
        
        self.bind(size=self.update_rect, pos=self.update_rect)

        self.label = Label(text="Enter Name and scores",
                           size_hint_y=None, height=50, font_size='20sp')
        
        self.text_inputs = [TextInput(hint_text=hint_text, size_hint_y=None, height=50, font_size='20sp') 
                            for hint_text in ["Enter your name of the child", "Weekly Assessment", 
                                              "Monthly Assessment", "Midterm", "End of Term"]]
        
        self.button = Button(text="Enter Data",
                             size_hint_y=None,size_hint_x=None, width=200,height=40, font_size='20sp')
        self.button.bind(on_press=self.save_to_excel)
        
        self.add_widget(self.label)
        for text_input in self.text_inputs:
            self.add_widget(text_input)
        self.add_widget(self.button)
        
    def save_to_excel(self, instance):
        try:
            wb = load_workbook("student_data.xlsx")
        except FileNotFoundError:
            wb = Workbook()
        ws = wb.active

        # Check if headers are already written

        # Get the index of the next available row
        next_row = ws.max_row + 1

    # Convert input data to appropriate data types
        data = [self.text_inputs[0].text]  # Name (string)
        for i in range(1, len(self.text_inputs)):
         try:
            data.append(float(self.text_inputs[i].text))  # Convert to float
         except ValueError:
            data.append(self.text_inputs[i].text) 
        ws.append(data)
        # Save the Excel file
        wb.save("student_data.xlsx")

        self.label.text = "Data saved to Excel sheet!"

    def update_rect(self, *args):
        self.rect.size = self.size
        self.rect.pos = self.pos

class Assessments(App):
    def build(self):
        return MyWidgetContainer()

if __name__ == "__main__":
    Assessments().run()
